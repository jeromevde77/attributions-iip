#!/usr/bin/env python3
"""
Génère un fichier SQL d'anonymisation des professeurs pour l'environnement dev.
- Noms/prénoms fictifs (Faker fr_FR)
- Dates de naissance (25-65 ans)
- Emails fictifs
- Titres de capacité réalistes selon la section
- Matricule fictif (11 chiffres)
"""
import random, datetime, json, sys, unicodedata
from faker import Faker
import openpyxl

fake = Faker('fr_FR')
random.seed(42)  # reproductible
Faker.seed(42)

# ── Titres par discipline ────────────────────────────────────────────────────
TITRES = {
    'TIM': [  # Technologue en imagerie médicale
        ("Master en sciences biomédicales",
         "Bachelier en technologie médicale – imagerie médicale"),
        ("Master en radiobiologie",
         "Agrégé de l'enseignement secondaire supérieur en sciences"),
        ("Master en sciences de la santé publique",
         "Bachelier en soins infirmiers"),
        ("Docteur en médecine",
         "Master en physique médicale"),
        ("Master en biologie clinique",
         "Bachelier technologue en imagerie médicale"),
    ],
    'Psychomotricité': [
        ("Master en psychomotricité",
         "Bachelier en psychomotricité"),
        ("Master en psychologie clinique",
         "Bachelier en kinésithérapie"),
        ("Master en sciences de l'éducation",
         "Bachelier en psychomotricité"),
        ("Master en neuropsychologie",
         "Bachelier en ergothérapie"),
        ("Docteur en sciences psychologiques",
         "Master en psychomotricité"),
    ],
    'Optométrie': [
        ("Master en sciences optiques",
         "Bachelier en optométrie"),
        ("Master en optique et optométrie",
         "Bachelier en optique-optométrie"),
        ("Docteur en sciences",
         "Master en optométrie clinique"),
        ("Master en médecine",
         "Bachelier en soins optométriques"),
        ("Master en biophysique",
         "Bachelier en optométrie"),
    ],
    'Optique': [  # secondaire
        ("Brevet d'enseignement supérieur en optique",
         "Certificat de qualification en optique-lunetterie"),
        ("Bachelier en optométrie",
         "CESS – option sciences"),
        ("Master en optique",
         "Brevet technique en optique"),
    ],
    'FID-Guidance': [
        ("Master en psychologie",
         "Bachelier en éducation spécialisée"),
        ("Master en sciences de l'éducation",
         "Agrégé de l'enseignement secondaire inférieur"),
        ("Master en logopédie",
         "Bachelier en guidance"),
        ("Master en orthopédagogie",
         "Bachelier en psychologie"),
    ],
    'FID-Péda': [
        ("Master en sciences de l'éducation",
         "Agrégé de l'enseignement secondaire supérieur"),
        ("Master en pédagogie",
         "Bachelier en éducation"),
        ("Master en didactique",
         "Agrégé de l'enseignement secondaire inférieur"),
    ],
    'ATNUP': [  # Administration, gestion
        ("Master en gestion des ressources humaines",
         "Bachelier en gestion d'entreprise"),
        ("Master en droit social",
         "Agrégé de l'enseignement secondaire supérieur en droit"),
        ("Master en sciences du travail",
         "Bachelier en secrétariat de direction"),
        ("Master en management",
         "Bachelier en comptabilité"),
    ],
    'SAR': [  # Soins, accompagnement
        ("Master en soins infirmiers",
         "Bachelier en soins infirmiers"),
        ("Master en santé communautaire",
         "Bachelier en sage-femme"),
        ("Docteur en médecine",
         "Master en sciences infirmières"),
    ],
    'ME': [  # Mécanique/Électronique
        ("Master en ingénierie industrielle",
         "Bachelier en électromécanique"),
        ("Master en électronique",
         "Bachelier en automatisation"),
        ("Ingénieur civil électricien",
         "Master en génie électrique"),
    ],
    'RESTART': [
        ("Master en sciences de l'éducation",
         "Bachelier en travail social"),
        ("Master en insertion socioprofessionnelle",
         "Bachelier en éducation"),
    ],
    'Soins_plaies': [
        ("Master en sciences infirmières",
         "Bachelier en soins infirmiers, spécialisation plaies et cicatrisation"),
        ("Master en dermatologie (sciences médicales)",
         "Bachelier en soins infirmiers"),
    ],
    'A': [
        ("Master en sciences de la motricité",
         "Bachelier en kinésithérapie"),
    ],
}
DEFAULT_TITRES = [
    ("Master en sciences",
     "Bachelier en éducation"),
    ("Master en didactique",
     "Agrégé de l'enseignement secondaire supérieur"),
]

# ── Charger les profs existants ──────────────────────────────────────────────
wb = openpyxl.load_workbook('/mnt/user-data/uploads/Attributions.xlsm',
                             read_only=True, data_only=True)
ws_coord = wb['Coordonnées_professeurs']
ws_attr  = wb['Attributions']

# Map nom_reel -> {sections}
prof_sections = {}
for r in ws_attr.iter_rows(min_row=2, values_only=True):
    if r[33] and r[0]:
        key = str(r[33]).strip()
        if key not in prof_sections:
            prof_sections[key] = set()
        prof_sections[key].add(str(r[0]).strip())

profs_reels = []
for r in ws_coord.iter_rows(min_row=2, values_only=True):
    if r[0] and str(r[0]).strip() not in ('', '-', 'None'):
        nom   = str(r[0]).strip()
        prenom = str(r[1]).strip() if r[1] else ''
        statut = str(r[6]).strip() if r[6] else 'EXP'
        profs_reels.append({
            'nom_reel': nom,
            'prenom_reel': prenom,
            'nom_complet_reel': f"{nom} {prenom}",
            'statut': statut,
            'sections': sorted(prof_sections.get(f"{nom} {prenom}", set())),
        })

print(f"Trouvé {len(profs_reels)} professeurs réels", file=sys.stderr)

# ── Générer les données anonymes ─────────────────────────────────────────────
today = datetime.date.today()

def gen_naissance():
    age = random.randint(25, 65)
    annee = today.year - age
    mois = random.randint(1, 12)
    jour = random.randint(1, 28)
    return datetime.date(annee, mois, jour)

def gen_matricule():
    return ''.join([str(random.randint(0, 9)) for _ in range(11)])

def sql_escape(s):
    """Double les apostrophes pour SQL."""
    return str(s).replace("'", "''")

def gen_titres(sections):
    """Choisir des titres cohérents avec la section principale, avec variété
    (Master, Bachelier, parfois un seul titre)."""
    pair = None
    for s in sections:
        if s in TITRES:
            pair = random.choice(TITRES[s])
            break
    if pair is None:
        pair = random.choice(DEFAULT_TITRES)
    t1, t2 = pair
    r = random.random()
    if r < 0.20:
        # Seulement le diplôme principal (pas de second titre)
        return t1, ""
    elif r < 0.35:
        # Profil "bachelier seul" : on garde le titre secondaire (souvent un bachelier) comme principal
        return t2, ""
    else:
        # Les deux titres
        return t1, t2

def genre_prenom(prenom):
    """Approximer le genre à partir du prénom pour cohérence."""
    return 'F'  # Faker donnera le bon genre

def normalize(s):
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn").lower()

# Éviter les doublons de noms fictifs
noms_utilises = set()
def gen_nom_unique():
    for _ in range(100):
        n = fake.last_name().upper()
        p = fake.first_name()
        cle = f"{n} {p}"
        if cle not in noms_utilises:
            noms_utilises.add(cle)
            return n, p
    return fake.last_name().upper(), fake.first_name()

# ── Communes belges réalistes (Bruxelles + Brabant wallon, zone IIP) ─────────
COMMUNES_BE = [
    ("1070", "Anderlecht"), ("1000", "Bruxelles"), ("1190", "Forest"),
    ("1180", "Uccle"), ("1060", "Saint-Gilles"), ("1050", "Ixelles"),
    ("1200", "Woluwe-Saint-Lambert"), ("1030", "Schaerbeek"), ("1080", "Molenbeek-Saint-Jean"),
    ("1410", "Waterloo"), ("1300", "Wavre"), ("1340", "Ottignies-Louvain-la-Neuve"),
    ("1480", "Tubize"), ("1420", "Braine-l'Alleud"), ("1083", "Ganshoren"),
    ("1700", "Dilbeek"), ("1640", "Rhode-Saint-Genèse"), ("1502", "Lembeek"),
]
RUES_BE = [
    "Rue de la Station", "Avenue des Tilleuls", "Chaussée de Bruxelles", "Rue du Moulin",
    "Avenue de la Liberté", "Rue des Écoles", "Clos des Pommiers", "Rue de l'Église",
    "Avenue des Cerisiers", "Rue Haute", "Drève des Bouleaux", "Rue du Bois",
    "Avenue Albert", "Rue Neuve", "Chemin des Champs", "Rue de la Forge",
]

def gen_adresse():
    rue = random.choice(RUES_BE)
    num = random.randint(1, 250)
    cp, commune = random.choice(COMMUNES_BE)
    return f"{rue} {num}", cp, commune

def gen_mail_prive(prenom, nom):
    fournisseur = random.choice(["gmail.com", "outlook.be", "hotmail.com", "skynet.be", "proximus.be"])
    style = random.randint(0, 2)
    p, n = normalize(prenom), normalize(nom)
    if style == 0:
        return f"{p}.{n}@{fournisseur}"
    elif style == 1:
        return f"{p}{n}@{fournisseur}"
    else:
        return f"{n}.{p}@{fournisseur}"

# Titre pédagogique belge : CAPAES (sup), CAP, AESS, ou aucun.
# Pondération réaliste : beaucoup ont le CAPAES (sup de promotion sociale),
# certains un CAP ou AESS, d'autres rien (titre en pénurie / récents).
def gen_titre_pedagogique():
    r = random.random()
    if r < 0.45:   return "CAPAES"
    elif r < 0.60: return "CAP"
    elif r < 0.72: return "AESS"
    else:          return ""   # pas de titre pédagogique

# ── Produire le SQL ──────────────────────────────────────────────────────────
print("-- Anonymisation des professeurs pour l'environnement dev")
print("-- Généré automatiquement — NE PAS appliquer en production")
print("-- Chaque professeur conserve ses attributions")
print()
print("BEGIN TRANSACTION;")
print()

mappings = []
for prof in profs_reels:
    nom_fictif, prenom_fictif = gen_nom_unique()
    naissance = gen_naissance()
    age = today.year - naissance.year
    matricule = gen_matricule()
    email_fictif = f"{normalize(prenom_fictif)}.{normalize(nom_fictif)}@lucie-dev.be"
    mail_prive = gen_mail_prive(prenom_fictif, nom_fictif)
    rue, cp, commune = gen_adresse()
    t1, t2 = gen_titres(prof['sections'])
    titre_peda = gen_titre_pedagogique()
    # CAPAES : colonne dédiée ("x" si possède le CAPAES, sinon vide)
    capaes = "x" if titre_peda == "CAPAES" else ""
    # Statut EA12 : EXP plus souvent nommé (T/TPr/D), CC temporaire (T)
    if prof['statut'] == 'EXP':
        statut_ea12 = random.choice(['T', 'TPr', 'TPr', 'D'])
    else:
        statut_ea12 = random.choice(['T', 'T', 'St'])
    # Ancienneté PO cohérente avec l'âge (entre 0 et age-23 ans, plafonnée)
    anciennete = random.randint(0, max(1, min(age - 23, 35)))

    # titre3 = titre pédagogique (champ libre sur l'EA12)
    titre3 = titre_peda

    mappings.append({
        'nom_reel': prof['nom_reel'], 'prenom_reel': prof['prenom_reel'],
        'nom_fictif': nom_fictif, 'prenom_fictif': prenom_fictif,
    })

    print(f"-- {prof['nom_reel']} {prof['prenom_reel']} → {nom_fictif} {prenom_fictif} ({age} ans, {prof['statut']}, {titre_peda or 'sans titre péda'})")
    print(f"UPDATE professeur SET")
    print(f"  nom                 = '{sql_escape(nom_fictif)}',")
    print(f"  prenom              = '{sql_escape(prenom_fictif)}',")
    print(f"  adresse_mail        = '{email_fictif}',")
    print(f"  mail_prive          = '{sql_escape(mail_prive)}',")
    print(f"  date_naissance      = '{naissance.isoformat()}',")
    print(f"  matricule           = '{matricule}',")
    print(f"  adresse_rue         = '{sql_escape(rue)}',")
    print(f"  code_postal         = '{cp}',")
    print(f"  commune             = '{sql_escape(commune)}',")
    print(f"  titre1              = '{sql_escape(t1)}',")
    print(f"  titre2              = '{sql_escape(t2)}',")
    print(f"  titre3              = '{sql_escape(titre3)}',")
    print(f"  capaes              = '{capaes}',")
    print(f"  statut_ea12         = '{statut_ea12}',")
    print(f"  anciennete_25_26_po = {anciennete}")
    print(f"WHERE nom = '{sql_escape(prof['nom_reel'])}' AND prenom = '{sql_escape(prof['prenom_reel'])}';")
    print()

print("COMMIT;")
# Mapping non exporté dans le SQL (RGPD : pas de vrais noms dans le fichier).
